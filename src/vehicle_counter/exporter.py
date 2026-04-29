import csv
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from . import config

DIRECTION_NEGATIVE_TO_POSITIVE = "negative_to_positive"
DIRECTION_POSITIVE_TO_NEGATIVE = "positive_to_negative"
DIRECTION_TO_EXPORT_LABEL = {
    DIRECTION_NEGATIVE_TO_POSITIVE: "A -> B",
    DIRECTION_POSITIVE_TO_NEGATIVE: "B -> A",
}
DEFAULT_LINE_IDS = ("line_1", "line_2", "line_3")
DEFAULT_CLASSES = config.SUPPORTED_COUNT_CLASSES
INTERVAL_MINUTES_OPTIONS = (1, 5, 15)


def export_results_file(file_path, results, source_details, direction_labels, settings):
    export_data = build_export_tables(results, source_details, direction_labels, settings)
    suffix = Path(file_path).suffix.lower()

    if suffix == ".csv":
        return write_csv_exports(file_path, export_data)

    if suffix == ".xlsx":
        return write_xlsx_export(file_path, export_data)

    return {
        "error": "Unsupported export format. Please choose a .csv or .xlsx file."
    }


def build_export_tables(results, source_details, direction_labels, settings):
    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    source_details = source_details or {}
    settings = settings or {}
    direction_labels = direction_labels or {}

    source_display_name = source_details.get("display_name") or get_source_label(
        source_details.get("original_input")
    )
    overall_counts = results.get("counts", {})

    summary_row = {
        "source_display_name": source_display_name or "Unknown",
        "source_kind": source_details.get("source_kind") or "unknown",
        "source_input": source_details.get("original_input") or "",
        "source_is_live": source_details.get("is_live", False),
        "source_stream_format": source_details.get("stream_format") or "",
        "exported_at": exported_at,
        "counting_mode": settings.get("counting_mode", config.DEFAULT_COUNTING_MODE),
        "active_preset": settings.get("preset_name", ""),
        "model_size": settings.get("model_size", ""),
        "model_size_label": settings.get("model_size_label", settings.get("model_size", "")),
        "confidence_threshold": settings.get("confidence_threshold", ""),
        "frame_skip": settings.get("frame_skip", ""),
        "imgsz": settings.get("imgsz", ""),
        "device": settings.get("device_label", settings.get("device", "")),
        "half_precision": settings.get("half", False),
        "preview_render_mode": settings.get("preview_render_mode", ""),
        "low_latency_mode": settings.get("prioritize_low_latency_live_streams", False),
        "annotated_video_enabled": settings.get("annotated_video_enabled", False),
        "enabled_classes": ", ".join(settings.get("enabled_classes", [])),
        "processed_frames": results.get("processed_frames", 0),
        "total_crossings": results.get("total", 0),
    }
    for class_name in DEFAULT_CLASSES:
        summary_row[f"overall_{class_name}"] = overall_counts.get(class_name, 0)

    line_summary_rows = []
    detail_rows = []

    line_results = results.get("line_results", {})
    for line_id in DEFAULT_LINE_IDS:
        line_result = line_results.get(line_id, {})
        labels = direction_labels.get(line_id, {})
        direction_a_label = labels.get(DIRECTION_NEGATIVE_TO_POSITIVE, "A -> B")
        direction_b_label = labels.get(DIRECTION_POSITIVE_TO_NEGATIVE, "B -> A")
        line_name = build_line_name(line_id, direction_a_label, direction_b_label)
        direction_counts = line_result.get("direction_counts", {})
        direction_a_total = sum(
            direction_counts.get(DIRECTION_NEGATIVE_TO_POSITIVE, {}).values()
        )
        direction_b_total = sum(
            direction_counts.get(DIRECTION_POSITIVE_TO_NEGATIVE, {}).values()
        )

        line_summary_rows.append(
            {
                "line_id": line_id,
                "line_name": line_name,
                "direction_a_label": direction_a_label,
                "direction_b_label": direction_b_label,
                "total_crossings": line_result.get("total", 0),
                "direction_a_total": direction_a_total,
                "direction_b_total": direction_b_total,
            }
        )

        for direction_key in (
            DIRECTION_NEGATIVE_TO_POSITIVE,
            DIRECTION_POSITIVE_TO_NEGATIVE,
        ):
            direction_name = DIRECTION_TO_EXPORT_LABEL[direction_key]
            movement_label = (
                direction_a_label
                if direction_key == DIRECTION_NEGATIVE_TO_POSITIVE
                else direction_b_label
            )
            class_counts = direction_counts.get(direction_key, {})
            for count_class in DEFAULT_CLASSES:
                detail_rows.append(
                    {
                        "line_id": line_id,
                        "line_name": line_name,
                        "direction": direction_name,
                        "direction_label": movement_label,
                        "count_class": count_class,
                        "count": class_counts.get(count_class, 0),
                    }
                )

    event_rows = build_event_rows(
        results=results,
        source_display_name=summary_row["source_display_name"],
        direction_labels=direction_labels,
    )

    return {
        "summary": [summary_row],
        "line_summary": line_summary_rows,
        "detail": detail_rows,
        "events": event_rows,
        "intervals": {
            interval_minutes: build_interval_rows(event_rows, interval_minutes)
            for interval_minutes in INTERVAL_MINUTES_OPTIONS
        },
    }


def write_csv_exports(file_path, export_data):
    base_path = Path(file_path)
    detail_path = base_path
    summary_path = base_path.with_name(f"{base_path.stem}_summary.csv")
    line_summary_path = base_path.with_name(f"{base_path.stem}_line_summary.csv")

    write_csv_table(detail_path, export_data["detail"])
    write_csv_table(summary_path, export_data["summary"])
    write_csv_table(line_summary_path, export_data["line_summary"])
    events_path = base_path.with_name(f"{base_path.stem}_events.csv")
    write_csv_table(events_path, export_data["events"])
    interval_paths = []
    for interval_minutes in INTERVAL_MINUTES_OPTIONS:
        interval_path = base_path.with_name(
            f"{base_path.stem}_interval_{interval_minutes}min.csv"
        )
        write_csv_table(interval_path, export_data["intervals"][interval_minutes])
        interval_paths.append(str(interval_path))

    return {
        "error": None,
        "primary_file": str(detail_path),
        "created_files": [
            str(detail_path),
            str(summary_path),
            str(line_summary_path),
            str(events_path),
            *interval_paths,
        ],
        "format": "csv",
    }


def write_xlsx_export(file_path, export_data):
    from openpyxl import Workbook

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    append_table(summary_sheet, export_data["summary"])

    line_summary_sheet = workbook.create_sheet("Line Summary")
    append_table(line_summary_sheet, export_data["line_summary"])

    detail_sheet = workbook.create_sheet("Detail")
    append_table(detail_sheet, export_data["detail"])

    events_sheet = workbook.create_sheet("Events")
    append_table(events_sheet, export_data["events"])

    for interval_minutes in INTERVAL_MINUTES_OPTIONS:
        interval_sheet = workbook.create_sheet(f"Interval {interval_minutes} Min")
        append_table(interval_sheet, export_data["intervals"][interval_minutes])

    workbook.save(file_path)
    return {
        "error": None,
        "primary_file": str(file_path),
        "created_files": [str(file_path)],
        "format": "xlsx",
    }


def write_csv_table(file_path, rows):
    with open(file_path, "w", newline="", encoding="utf-8-sig") as csv_file:
        if not rows:
            csv_file.write("")
            return
        writer = csv.DictWriter(csv_file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def append_table(sheet, rows):
    if not rows:
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def build_line_name(line_id, direction_a_label, direction_b_label):
    display_name = line_id.replace("_", " ").title()
    if (
        direction_a_label == DIRECTION_TO_EXPORT_LABEL[DIRECTION_NEGATIVE_TO_POSITIVE]
        and direction_b_label == DIRECTION_TO_EXPORT_LABEL[DIRECTION_POSITIVE_TO_NEGATIVE]
    ):
        return display_name
    return f"{display_name} ({direction_a_label} / {direction_b_label})"


def build_event_rows(results, source_display_name, direction_labels):
    event_rows = []
    for event in results.get("events", []):
        line_id = event.get("line_id", "")
        labels = direction_labels.get(line_id, {})
        direction = event.get("direction", "")
        count_class = event.get("count_class", event.get("vehicle_class", ""))
        direction_label = labels.get(direction, DIRECTION_TO_EXPORT_LABEL.get(direction, ""))
        direction_a_label = labels.get(DIRECTION_NEGATIVE_TO_POSITIVE, "A -> B")
        direction_b_label = labels.get(DIRECTION_POSITIVE_TO_NEGATIVE, "B -> A")
        event_rows.append(
            {
                "event_id": event.get("event_id", ""),
                "source_display_name": source_display_name,
                "line_id": line_id,
                "line_name": build_line_name(line_id, direction_a_label, direction_b_label),
                "direction": DIRECTION_TO_EXPORT_LABEL.get(direction, direction),
                "direction_label": direction_label,
                "count_class": count_class,
                "track_id": event.get("track_id", ""),
                "frame_index": event.get("frame_index", ""),
                "elapsed_seconds": event.get("elapsed_seconds", ""),
                "event_timecode": event.get("event_timecode", ""),
            }
        )
    return event_rows


def build_interval_rows(event_rows, interval_minutes):
    interval_seconds = interval_minutes * 60
    grouped_counts = {}

    for event in event_rows:
        elapsed_seconds = event.get("elapsed_seconds")
        if elapsed_seconds in ("", None):
            continue

        interval_index = int(float(elapsed_seconds) // interval_seconds)
        interval_start_seconds = interval_index * interval_seconds
        interval_end_seconds = interval_start_seconds + interval_seconds
        group_key = (
            interval_index,
            event.get("line_id", ""),
            event.get("line_name", ""),
            event.get("direction", ""),
            event.get("direction_label", ""),
            event.get("count_class", ""),
        )
        grouped_counts[group_key] = grouped_counts.get(group_key, 0) + 1

    interval_rows = []
    for group_key in sorted(grouped_counts.keys()):
        (
            interval_index,
            line_id,
            line_name,
            direction,
            direction_label,
            count_class,
        ) = group_key
        interval_start_seconds = interval_index * interval_seconds
        interval_end_seconds = interval_start_seconds + interval_seconds
        interval_rows.append(
            {
                "interval_minutes": interval_minutes,
                "interval_index": interval_index,
                "interval_start_seconds": interval_start_seconds,
                "interval_end_seconds": interval_end_seconds,
                "interval_start_timecode": format_interval_timecode(interval_start_seconds),
                "interval_end_timecode": format_interval_timecode(interval_end_seconds),
                "line_id": line_id,
                "line_name": line_name,
                "direction": direction,
                "direction_label": direction_label,
                "count_class": count_class,
                "count": grouped_counts[group_key],
            }
        )

    return interval_rows


def format_interval_timecode(total_seconds):
    total_seconds = max(0.0, float(total_seconds))
    total_milliseconds = round(total_seconds * 1000)
    hours = total_milliseconds // 3600000
    remaining = total_milliseconds % 3600000
    minutes = remaining // 60000
    remaining %= 60000
    seconds = remaining // 1000
    milliseconds = remaining % 1000
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"


def get_source_label(source_input):
    if not source_input:
        return None

    parsed = urlparse(source_input)
    if parsed.scheme in {"http", "https"}:
        return source_input

    return Path(source_input).name
